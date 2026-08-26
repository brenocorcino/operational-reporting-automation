from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .transform import ReportData

NAVY = "17365D"
BLUE = "2F75B5"
LIGHT_BLUE = "D9EAF7"
GREEN = "E2F0D9"
RED = "FCE4D6"
WHITE = "FFFFFF"


def _write_frame(writer: pd.ExcelWriter, name: str, frame: pd.DataFrame) -> None:
    frame.to_excel(writer, sheet_name=name, index=False)


def _style_table(worksheet) -> None:
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for cell in worksheet[1]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center")

    for column in range(1, worksheet.max_column + 1):
        values = [
            worksheet.cell(row, column).value for row in range(1, worksheet.max_row + 1)
        ]
        width = min(36, max(12, max(len(str(value or "")) for value in values) + 2))
        worksheet.column_dimensions[get_column_letter(column)].width = width


def build_workbook(data: ReportData, output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        pd.DataFrame([data.summary]).to_excel(
            writer, sheet_name="Summary Data", index=False
        )
        _write_frame(writer, "Collection", data.collection)
        _write_frame(writer, "Transfers", data.transfer)
        _write_frame(writer, "Exceptions", data.exceptions)

    workbook = load_workbook(output_path)
    summary_data = workbook["Summary Data"]
    summary_data.sheet_state = "hidden"
    summary = workbook.create_sheet("Summary", 0)

    summary.merge_cells("A1:F1")
    summary["A1"] = "Operational Performance Report"
    summary["A1"].fill = PatternFill("solid", fgColor=NAVY)
    summary["A1"].font = Font(color=WHITE, bold=True, size=18)
    summary["A1"].alignment = Alignment(horizontal="center")
    summary.row_dimensions[1].height = 30

    cards = [
        ("A3", "Report date", data.summary["report_date"], LIGHT_BLUE),
        ("C3", "Planned packages", data.summary["planned_packages"], LIGHT_BLUE),
        ("E3", "Collected packages", data.summary["collected_packages"], GREEN),
        ("A6", "Collection rate", data.summary["collection_rate"], GREEN),
        ("C6", "Transfer on-time rate", data.summary["transfer_on_time_rate"], GREEN),
        ("E6", "Affected shipments", data.summary["affected_shipments"], RED),
    ]

    for anchor, label, value, color in cards:
        cell = summary[anchor]
        cell.value = label
        cell.font = Font(bold=True, color=NAVY)
        cell.fill = PatternFill("solid", fgColor=color)
        value_cell = summary.cell(cell.row + 1, cell.column)
        value_cell.value = value
        value_cell.font = Font(bold=True, size=16)
        value_cell.fill = PatternFill("solid", fgColor=color)
        if "rate" in label.lower():
            value_cell.number_format = "0.00%"

    summary.column_dimensions["A"].width = 23
    summary.column_dimensions["B"].width = 4
    summary.column_dimensions["C"].width = 23
    summary.column_dimensions["D"].width = 4
    summary.column_dimensions["E"].width = 23
    summary.column_dimensions["F"].width = 4

    collection_sheet = workbook["Collection"]
    rate_column = next(
        cell.column for cell in collection_sheet[1] if cell.value == "collection_rate"
    )
    for row in range(2, collection_sheet.max_row + 1):
        collection_sheet.cell(row, rate_column).number_format = "0.00%"

    transfer_sheet = workbook["Transfers"]
    transfer_rate_column = next(
        cell.column for cell in transfer_sheet[1] if cell.value == "on_time_rate"
    )
    for row in range(2, transfer_sheet.max_row + 1):
        transfer_sheet.cell(row, transfer_rate_column).number_format = "0.00%"

    for sheet_name in ("Collection", "Transfers", "Exceptions"):
        _style_table(workbook[sheet_name])

    chart = BarChart()
    chart.title = "Collected packages by hub"
    chart.y_axis.title = "Packages"
    chart.x_axis.title = "Hub"
    collected_column = next(
        cell.column
        for cell in collection_sheet[1]
        if cell.value == "collected_packages"
    )
    hub_column = next(
        cell.column for cell in collection_sheet[1] if cell.value == "hub"
    )
    chart.add_data(
        Reference(
            collection_sheet,
            min_col=collected_column,
            min_row=1,
            max_row=collection_sheet.max_row,
        ),
        titles_from_data=True,
    )
    chart.set_categories(
        Reference(
            collection_sheet,
            min_col=hub_column,
            min_row=2,
            max_row=collection_sheet.max_row,
        )
    )
    chart.height = 7
    chart.width = 15
    summary.add_chart(chart, "A10")

    workbook.save(output_path)
    return output_path

from __future__ import annotations

import tempfile
import unittest
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from report_automation.transform import prepare_reports
from report_automation.workbook import build_workbook


class WorkbookTests(unittest.TestCase):
    def test_builds_formatted_workbook_with_expected_sheets(self) -> None:
        project = Path(__file__).resolve().parents[1]
        report_date = date(2026, 8, 25)

        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / report_date.isoformat()
            source.mkdir()
            mapping = {
                "collection": "collection_performance.csv",
                "transfer": "transfer_performance.csv",
                "exceptions": "scan_exceptions.csv",
            }
            for target, source_name in mapping.items():
                (source / f"{target}_{report_date.isoformat()}.csv").write_bytes(
                    (project / "data" / "source" / source_name).read_bytes()
                )

            data = prepare_reports(source, report_date)
            output = Path(temp_dir) / "report.xlsx"
            build_workbook(data, output)
            workbook = load_workbook(output)

            self.assertEqual(
                workbook.sheetnames,
                ["Summary", "Summary Data", "Collection", "Transfers", "Exceptions"],
            )
            self.assertEqual(
                workbook["Summary"]["A1"].value, "Operational Performance Report"
            )
            self.assertEqual(workbook["Summary Data"].sheet_state, "hidden")
            self.assertEqual(len(workbook["Summary"]._charts), 1)
            self.assertEqual(workbook["Collection"].freeze_panes, "A2")


if __name__ == "__main__":
    unittest.main()
